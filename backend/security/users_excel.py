from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple


@dataclass(frozen=True)
class ExcelUsers:
    users_by_email: Dict[str, str]  # email -> role
    local_users: Dict[str, dict]  # username -> {code, role, display_name, email?}


_CACHE: Tuple[Optional[str], float, ExcelUsers] = (None, -1.0, ExcelUsers(users_by_email={}, local_users={}))


def invalidate_users_excel_cache() -> None:
    """Force le prochain chargement à relire le fichier Excel (après édition du rôle)."""
    global _CACHE
    _CACHE = (None, -1.0, ExcelUsers(users_by_email={}, local_users={}))


def _norm_email(email: str) -> str:
    return (email or "").strip().lower()

def _norm_username(username: str) -> str:
    return (username or "").strip().lower()


def load_users_from_excel(path: str) -> ExcelUsers:
    """
    Load a simple directory of users from an .xlsx file.

    Expected columns (header row, case-insensitive):
    - email (required)
    - role (required)  e.g. admin | analyst_2025 | analyst_2026
    - enabled (optional)  values like 1/true/yes/oui to enable; 0/false/no/non to disable

    The file is cached by mtime, so updates apply automatically without restarting the server.
    """
    global _CACHE
    p = (path or "").strip()
    if not p:
        return ExcelUsers(users_by_email={}, local_users={})

    try:
        fp = str(Path(p).expanduser())
        stat = os.stat(fp)
        mtime = float(stat.st_mtime)
    except Exception:
        return ExcelUsers(users_by_email={}, local_users={})

    cache_path, cache_mtime, cache_data = _CACHE
    if cache_path == fp and cache_mtime == mtime:
        return cache_data

    users: Dict[str, str] = {}
    local_users: Dict[str, dict] = {}
    try:
        # openpyxl is already in requirements.txt
        from openpyxl import load_workbook

        wb = load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            data = ExcelUsers(users_by_email={}, local_users={})
            _CACHE = (fp, mtime, data)
            return data

        cols = {str(v).strip().lower(): i for i, v in enumerate(header) if v is not None and str(v).strip()}
        # Two supported modes (same sheet):
        # - SSO mapping: email + role (+ enabled)
        # - Local code login: username + code + role (+ display_name/email/enabled)
        has_email_role = ("email" in cols and "role" in cols)
        has_local_code = ("username" in cols and "code" in cols and "role" in cols)
        if not has_email_role and not has_local_code:
            data = ExcelUsers(users_by_email={}, local_users={})
            _CACHE = (fp, mtime, data)
            return data

        enabled_idx = cols.get("enabled")
        display_idx = cols.get("display_name")
        email_idx = cols.get("email")
        username_idx = cols.get("username")
        code_idx = cols.get("code")
        role_idx = cols.get("role")

        def is_enabled(val) -> bool:
            if val is None or val == "":
                return True
            s = str(val).strip().lower()
            if s in {"0", "false", "no", "non", "disabled", "off"}:
                return False
            if s in {"1", "true", "yes", "oui", "enabled", "on"}:
                return True
            # default: treat unknown as enabled
            return True

        for r in rows:
            if not r:
                continue
            try:
                if enabled_idx is not None:
                    val = r[enabled_idx] if enabled_idx < len(r) else None
                    if not is_enabled(val):
                        continue
                role = str(r[role_idx] if role_idx is not None and role_idx < len(r) else "").strip()
                if not role:
                    continue

                # SSO mapping by email
                if email_idx is not None:
                    email = _norm_email(r[email_idx] if email_idx < len(r) else "")
                    if email:
                        users[email] = role

                # Local mapping by username + code
                if username_idx is not None and code_idx is not None:
                    username = _norm_username(r[username_idx] if username_idx < len(r) else "")
                    code = str(r[code_idx] if code_idx < len(r) else "").strip()
                    if username and code:
                        display_name = str(r[display_idx] if display_idx is not None and display_idx < len(r) else "").strip()
                        local_users[username] = {
                            "username": username,
                            "code": code,
                            "role": role,
                            "display_name": display_name or username,
                            "email": _norm_email(r[email_idx] if email_idx is not None and email_idx < len(r) else "") if email_idx is not None else "",
                        }
            except Exception:
                continue
    except Exception:
        users = {}
        local_users = {}

    data = ExcelUsers(users_by_email=users, local_users=local_users)
    _CACHE = (fp, mtime, data)
    return data


def update_local_code_in_excel(path: str, *, username: str, new_code: str) -> bool:
    """
    Update the `code` for a local user in the Excel directory.
    Returns True if a row was updated and saved.
    """
    p = (path or "").strip()
    if not p:
        return False
    fp = str(Path(p).expanduser())
    u = _norm_username(username)
    if not u:
        return False
    code = str(new_code or "").strip()
    if not code:
        return False

    try:
        from openpyxl import load_workbook

        wb = load_workbook(fp, read_only=False, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=False)
        header = next(rows, None)
        if not header:
            return False
        cols = {str(c.value).strip().lower(): i for i, c in enumerate(header) if c.value is not None and str(c.value).strip()}
        if "username" not in cols or "code" not in cols:
            return False
        username_idx = cols["username"]
        code_idx = cols["code"]

        updated = False
        for r in rows:
            try:
                if username_idx >= len(r) or code_idx >= len(r):
                    continue
                cell_u = r[username_idx]
                cell_c = r[code_idx]
                if _norm_username(str(cell_u.value or "")) != u:
                    continue
                cell_c.value = code
                updated = True
                break
            except Exception:
                continue
        if not updated:
            return False

        wb.save(fp)
    except Exception:
        return False

    # Invalidate cache so next load reflects changes
    global _CACHE
    _CACHE = (None, -1.0, ExcelUsers(users_by_email={}, local_users={}))
    return True

